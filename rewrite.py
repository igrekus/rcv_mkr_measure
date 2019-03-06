import datetime
import time
import openpyxl
import pyvisa
import serial
import sys

from pnamock import PnaMock
from serialmock import SerialMock


is_mock = True
# ser = serial.Serial(port='COM10', baudrate=115200)
ser = SerialMock()
pna_mock = PnaMock()

bits = {
    5: [9, 10],
    6: [7, 8],
    3: [11, 12],
    4: [13, 14]
}

phases = [
    22.5,
    45.0,
    90.0,
    180.0
]

bit_states = {
    0: [0, 0, 0, 0],
    1: [1, 0, 0, 0],
    2: [0, 1, 0, 0],
    3: [1, 1, 0, 0],
    4: [0, 0, 1, 0],
    5: [1, 0, 1, 0],
    6: [0, 1, 1, 0],
    7: [1, 1, 1, 0],
    8: [0, 0, 0, 1],
    9: [1, 0, 0, 1],
    10: [0, 1, 0, 1],
    11: [1, 1, 0, 1],
    12: [0, 0, 1, 1],
    13: [1, 0, 1, 1],
    14: [0, 1, 1, 1],
    15: [1, 1, 1, 1]
}


def jerome_init(jerome):
    jerome.write(b'$KE,IO,SET,7,0\r\n')
    jerome.write(b'$KE,IO,SET,8,0\r\n')
    jerome.write(b'$KE,IO,SET,9,0\r\n')
    jerome.write(b'$KE,IO,SET,10,0\r\n')
    jerome.write(b'$KE,IO,SET,11,0\r\n')
    jerome.write(b'$KE,IO,SET,12,0\r\n')
    jerome.write(b'$KE,IO,SET,13,0\r\n')
    jerome.write(b'$KE,IO,SET,14,0\r\n')
    jerome.write(b'$KE,WRA,000000010101010000000000\r\n')


def jerome_set_bit_pattern(pattern: list, jerome):
    for bit, state in zip(bits.keys(), pattern):
        write_bit(bit, state, jerome)


def write_bit(bit_n, state, jerome):
    pos_addr, neg_addr = bits[bit_n]
    jerome.write(bytes(f'$KE,WR,{pos_addr},{state}', encoding='ascii'))
    jerome.write(bytes(f'$KE,WR,{neg_addr},{state ^ 1}', encoding='ascii'))


def pna_init(pna):
    pna.write('SYST:PRES')
    pna.query('*OPC?')
    pna.write('CALC:PAR:DEL:ALL')

    pna.write('DISP:WIND2 ON')

    pna.write('CALC1:PAR:DEF "CH1_S21",S21')
    pna.write('CALC2:PAR:DEF "CH2_S21",S21')
    pna.write('CALC1:PAR:DEF "CH1_S11",S11')
    pna.write('CALC1:PAR:DEF "CH1_S22",S22')

    pna.write('SENS1:CORR:CSET:ACT "-20dBm_1.1-1.4G",1')
    pna.write('SENS2:CORR:CSET:ACT "-20dBm_1.1-1.4G",1')

    pna.write('DISP:WIND1:TRAC1:FEED "CH1_S21"')
    pna.write('DISP:WIND2:TRAC1:FEED "CH2_S21"')
    pna.write('DISP:WIND1:TRAC2:FEED "CH1_S11"')
    pna.write('DISP:WIND1:TRAC3:FEED "CH1_S22"')

    pna.write('SENS1:SWE:MODE CONT')
    pna.write('SENS2:SWE:MODE CONT')

    pna.write('CALC1:FORM MLOG')
    pna.write('CALC2:FORM UPH')
    pna.write('DISP:WIND1:TRAC1:Y:SCAL:AUTO')
    pna.write('DISP:WIND2:TRAC1:Y:SCAL:AUTO')


def VSWR_calc(inp_S: list):
    tmp_inp_S = map(lambda x: x/20, inp_S)
    modS = list(map(lambda x: pow(10, x), tmp_inp_S))
    outp_G = map(lambda z: z[0] / z[1] if z[1] != 0 else 0.000001, zip(map(lambda x: 1 + x, modS), map(lambda x: 1 - x, modS)))

    return list(outp_G)


def init_file(file_path, freq, states, gamma_inp, gamma_outp, mS21, pS21):

    first_16_nonzero_elem_indices = [idx for idx, x in enumerate(states) if x != 0]

    num_states = len(first_16_nonzero_elem_indices) + 1

    N = len(freq)
    ofs = 2

    wb = openpyxl.Workbook()
    ws = wb.active

    str_y, str_mth, str_d, str_h, str_m, *rest = datetime.datetime.now().timetuple()
    str_time = f'{str_h}:{str_m}'
    str_date = f'{str_d}.{str_mth}.{str_y}'

    title = 'приемный модуль МШУ'
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=3, value=f'date:{str_date} time:{str_time}')
    ws.cell(row=3, column=1, value='пункт')
    ws.cell(row=3, column=2, value='частота, МГц')

    for j in range(1, num_states + 1):
        k = j * 4 - 3 + ofs
        ws.cell(row=3, column=k + 0, value='SWR_in')
        ws.cell(row=3, column=k + 1, value='SWR_out')
        ws.cell(row=3, column=k + 2, value='S21,дБ')
        ws.cell(row=3, column=k + 3, value='phase,гр.')

        ws.cell(row=4, column=k + 0, value=f'{states[j - 1]} гр.')
        ws.cell(row=4, column=k + 1, value=f'{states[j - 1]} гр.')
        ws.cell(row=4, column=k + 2, value=f'{states[j - 1]} гр.')
        ws.cell(row=4, column=k + 3, value=f'{states[j - 1]} гр.')

    for i in range(N):
        ws.cell(row=5 + i, column=1, value=f'{i:03d}')
        ws.cell(row=5 + i, column=2, value=freq[i] * 1e-6)

    pivot_row = 5
    pivot_col = 0

    for j in range(num_states):
        k = (j + 1) * 4 - 3 + ofs
        for i in range(N):
            ws.cell(row=i + pivot_row, column=k + 0 + pivot_col, value=gamma_inp[j][i])
            ws.cell(row=i + pivot_row, column=k + 1 + pivot_col, value=gamma_outp[j][i])
            ws.cell(row=i + pivot_row, column=k + 2 + pivot_col, value=mS21[j][i])
            ws.cell(row=i + pivot_row, column=k + 3 + pivot_col, value=pS21[j][i])

    wb.save(file_path)
    print('saved .xlsx:', file_path)


def save_file(file_path, freq, gamma_inp, gamma_outp, mS21, pS21, s21_min, s21_max, delta_s21, delta_Kp, states):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    xls_str_start = 6  # перва¤ строка для записи значений
    xls_cln_start = 3  # первый столбец для записи значений

    N = len(freq)
    temp = [0] * N

    first_16_nonzero_elem_indices = [idx for idx, x in enumerate(states) if x != 0]
    num_states = len(first_16_nonzero_elem_indices) + 1

    indx = 0  # индекс смещения по колонкам

    for j in range(1, num_states + 1):
        ptr_cln = indx + xls_cln_start
        temp[:] = gamma_inp[j - 1][:]
        for row, value in enumerate(temp):
            ws.cell(row=xls_str_start + row, column=ptr_cln, value=value)
        indx += 4

    indx = 1

    for j in range(1, num_states + 1):
        ptr_cln = indx + xls_cln_start
        temp[:] = gamma_outp(j - 1)[:]
        for row, value in enumerate(temp):
            ws.cell(row=xls_str_start + row, column=ptr_cln, value=value)
        indx += 4

    indx = 2

    for j in range(1, num_states + 1):
        ptr_cln = indx + xls_cln_start
        temp[:] = mS21[j - 1][:]
        for row, value in enumerate(temp):
            ws.cell(row=xls_str_start + row, column=ptr_cln, value=value)
        indx += 4

    indx = 3

    for j in range(1, num_states + 1):
        ptr_cln = indx + xls_cln_start
        temp[:] = pS21[j - 1][:]
        for row, value in enumerate(temp):
            ws.cell(row=xls_str_start + row, column=ptr_cln, value=value)
        indx += 4

    temp = [0] * 3
    indx = 0

    for j in range(1, num_states + 1):
        ptr_cln1 = indx + xls_cln_start
        ptr_cln2 = indx + xls_cln_start + 2
        temp[1] = s21_max[j]
        temp[2] = s21_min[j]
        temp[3] = delta_s21[j]
        for col, value in enumerate(temp):
            ws.cell(row=xls_str_start - 3, column=ptr_cln1 + col, value=value)
        indx += 4

    data2str = delta_Kp
    ptr_cln = 2
    ws.cell(row=xls_str_start - 3, column=ptr_cln, value=delta_Kp)

    wb.save(file_path)

    print('data saved to:', file_path)


def get_measurement(pna, calc, param):
    return pna.query(f'CALC{calc}:PAR:SEL "{param}";CALC{calc}:DATA? FDATA')


def get_freqs(pna):
    pna.write('CALC1:PAR:SEL "CH1_S21"')
    pna.write('FORM:DATA REAL,32')
    return pna.query('SENS1:X?')


def get_phase_value(pattern):
    return sum([v * p for v, p in zip(phases, pattern)])


def measure_s_params(pna, jerome):
    mag_s11_arr = list()
    mag_s22_arr = list()
    mag_s21_arr = list()
    phs_s21_arr = list()
    st_arr = list()
    for state in bit_states.values():
        st_arr.append(get_phase_value(state))  # phs_state

        jerome_set_bit_pattern(state, jerome=jerome)
        mag_s21_arr.append(get_measurement(pna=pna, calc=1, param='CH1_S21'))
        phs_s21_arr.append(get_measurement(pna=pna, calc=2, param='CH2_S21'))
        mag_s11_arr.append(get_measurement(pna=pna, calc=1, param='CH1_S11'))
        mag_s22_arr.append(get_measurement(pna=pna, calc=1, param='CH1_S22'))

    return mag_s21_arr, phs_s21_arr, mag_s11_arr, mag_s22_arr, st_arr


def reset_commutator(jerome):
    jerome_set_bit_pattern([1, 1, 1, 1], jerome=jerome)
    jerome.close()


def calc_gammas(mag_s11_arr, mag_s22_arr):
    gamma_inp = [VSWR_calc(dataset) for dataset in mag_s11_arr]
    gamma_outp = [VSWR_calc(dataset) for dataset in mag_s22_arr]
    return gamma_inp, gamma_outp


def find_freq_index(freqs: list, threshold):
    df = freqs[1] - freqs[0]
    return next((i for i, f in enumerate(freqs) if (threshold - df) < f < (threshold + df)), 0)


def calc_s21_stats(ind_dn_frq, ind_up_frq, mag_s21_arr):
    s21_max = list()
    s21_min = list()
    delta_s21 = list()

    for data in mag_s21_arr:
        temp = data[ind_dn_frq:ind_up_frq + 1]
        mx, mn = max(temp), min(temp)
        s21_max.append(mx)
        s21_min.append(mn)
        delta_s21.append(mx - mn)

    return s21_max, s21_min, delta_s21


def calc_out_stats(s21_max, s21_min):
    s21_MAX = max(s21_max)
    s21_MIN = min(s21_min)
    delta_Kp = abs(s21_MAX) - abs(s21_MIN)
    sred_Kp = (s21_MAX + s21_MIN) / 2
    return delta_Kp, s21_MAX, s21_MIN, sred_Kp


def ref_pts_stats(gamma_inp, ind_dn_frq, ind_up_frq, eps, threshold):
    return [[1 for g in gamma[ind_dn_frq:ind_up_frq + 1] if g > threshold + eps] for gamma in gamma_inp]


def measure(pna_addr='TCPIP0::192.168.1.61::inst0::INSTR'):

    file_name = '.\\xlsx\\out.xlsx'

    jerome, pna = prepare_rig(pna_addr)

    freqs = get_freqs(pna)

    mag_s21_arr, phs_s21_arr, mag_s11_arr, mag_s22_arr, st_arr = measure_s_params(pna, jerome)

    pna.close()
    reset_commutator(jerome)

    gamma_inp, gamma_outp = calc_gammas(mag_s11_arr, mag_s22_arr)

    # init_file(file_name, freqs, st_arr, gamma_inp, gamma_outp, mag_s21_arr, phs_s21_arr)

    ind_dn_frq = find_freq_index(freqs, threshold=1.21e9)
    ind_up_frq = find_freq_index(freqs, threshold=1.31e9)

    s21_max, s21_min, delta_s21 = calc_s21_stats(ind_dn_frq, ind_up_frq, mag_s21_arr)

    delta_Kp, s21_MAX, s21_MIN, sred_Kp = calc_out_stats(s21_max, s21_min)

    print('delta_Kp=', delta_Kp)
    print('approx median amp=', sred_Kp)
    print('Max_S21 = ', s21_MAX)
    print('Min_S21 = ', s21_MIN)

    ref_pnt_inp = ref_pts_stats(gamma_inp, ind_dn_frq, ind_up_frq, eps=1e-1, threshold=1.5)
    ref_pnt_outp = ref_pts_stats(gamma_outp, ind_dn_frq, ind_up_frq, eps=1e-1, threshold=1.5)

    summ_inp = [sum(pts) for pts in ref_pnt_inp]
    summ_outp = [sum(pts) for pts in ref_pnt_outp]

    if sum(summ_inp) == 0:
        print('WSVR in < 1.5')
    else:
        print('!!! WSVR in > 1.5 !!!')

    if sum(summ_outp) == 0:
        print('WSVR out < 1.5')
    else:
        print('!!! WS?VR out > 1.5 !!!')


def prepare_rig(pna_addr):
    jerome, pna = find_rig(pna_addr)

    if not jerome:
        print('error: jerome not found')
        sys.exit(1)
    if not pna:
        print('error: PNA not found')
        sys.exit(2)

    jerome_init(jerome)
    pna_init(pna)

    return jerome, pna


def find_rig(pna_addr):
    if is_mock:
        return SerialMock(), PnaMock()
    return find_jerome(), find_pna(pna_addr)


def find_jerome():

    def find_available_ports():
        available_ports = list()
        for i in range(256):
            port = f'COM{i+1}'
            try:
                s = serial.Serial(port=port, baudrate=115200)
                s.close()
                available_ports.append(port)
            except (OSError, serial.SerialException):
                pass
        return available_ports

    for port in find_available_ports():
        s = serial.Serial(port=port, baudrate=115200)
        s.write(b'$KE\r\n')
        time.sleep(0.1)
        ans = s.read_all()
        if b'#OK' in ans:
            return s

    return None


def find_pna(pna_addr):
    rm = pyvisa.ResourceManager()
    inst = rm.open_resource(pna_addr)

    ans = inst.query('*IDN?')
    if 'E8362B' in ans:
        return inst

    return None


if __name__ == '__main__':
    measure(pna_addr=sys.argv[1])
