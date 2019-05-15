import datetime
import subprocess
import openpyxl

from openpyxl.chart import LineChart


class ExcelResultFinal:

    def __init__(self, result):
        self._result = result

    @property
    def file_path(self):
        year, month, day, hour, minute, *rest = datetime.datetime.now().timetuple()
        return f'.\\xlsx\\VSWR-{year}-{month:02d}-{day:02d}-{hour:02d}.{minute:02d}-full.xlsx'

    def save(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        wb_header_cell = ws['A1']

        # for row in rows:
        #     ws.append(row)
        max_row = len(self._result._freqs) + 1

        freq_cell = wb_header_cell.offset(0, 22)
        freq_cell.value = 'freq'
        for i, freq in enumerate(self._result._freqs):
            freq_cell.offset(1 + i, 0).value = freq / 1_000_000_000

        s21_cell = freq_cell.offset(0, 1)
        for i, state in enumerate(self._result._states):
            s21_cell.offset(0, i).value = f's21:{state:.01f}°'
        for col, mags in enumerate(self._result._mag_s21s):
            for row, mag in enumerate(mags):
                s21_cell.offset(1 + row, col).value = mag

        swr_in_cell = s21_cell.offset(0, 16)
        for i, state in enumerate(self._result._states):
            swr_in_cell.offset(0, i).value = f'swr in:{state:.01f}°'
        for col, swrs in enumerate(self._result._gamma_input):
            for row, swr in enumerate(swrs):
                swr_in_cell.offset(1 + row, col).value = swr

        swr_out_cell = swr_in_cell.offset(0, 16)
        for i, state in enumerate(self._result._states):
            swr_out_cell.offset(0, i).value = f'swr out:{state:.01f}°'
        for col, swrs in enumerate(self._result._gamma_output):
            for row, swr in enumerate(swrs):
                swr_out_cell.offset(1 + row, col).value = swr

        phase_cell = swr_out_cell.offset(0, 16)
        for i, state in enumerate(self._result._states[1:]):
            phase_cell.offset(0, i).value = f'phs:{state:.01f}°'
        for col, phases in enumerate(self._result._phases):
            for row, phase in enumerate(phases):
                phase_cell.offset(1 + row, col).value = phase

        phase_err_cell = phase_cell.offset(0, 15)
        for i, state in enumerate(self._result._states[1:]):
            phase_err_cell.offset(0, i).value = f'p err:{state:.01f}°'
        for col, phase_errs in enumerate(self._result._phase_errs):
            for row, phase_err in enumerate(phase_errs):
                phase_err_cell.offset(1 + row, col).value = phase_err

        # TODO calculate cell range instead of hard-coded string
        cats = openpyxl.chart.Reference(ws, range_string=f'Sheet!$W$2:$W${max_row}')

        chart_s21 = openpyxl.chart.LineChart()
        chart_s21.style = 2
        chart_s21.x_axis.title = 'F, GHz'
        chart_s21.y_axis.title = 'S21'
        vals = openpyxl.chart.Reference(ws, range_string=f'Sheet!$X$1:$AM${max_row}')
        chart_s21.add_data(vals, titles_from_data=True)
        chart_s21.set_categories(cats)
        ws.add_chart(chart_s21, 'B2')

        chart_swr_in = openpyxl.chart.LineChart()
        chart_swr_in.style = 2
        chart_swr_in.x_axis.title = 'F, GHz'
        chart_swr_in.y_axis.title = 'SWR in, dB'
        vals = openpyxl.chart.Reference(ws, range_string=f'Sheet!$AN$1:$BC${max_row}')
        chart_swr_in.add_data(vals, titles_from_data=True)
        chart_swr_in.set_categories(cats)
        ws.add_chart(chart_swr_in, 'L2')

        chart_swr_out = openpyxl.chart.LineChart()
        chart_swr_out.style = 2
        chart_swr_out.x_axis.title = 'F, GHz'
        chart_swr_out.y_axis.title = 'SWR out, dB'
        vals = openpyxl.chart.Reference(ws, range_string=f'Sheet!$BD$1:$BS${max_row}')
        chart_swr_out.add_data(vals, titles_from_data=True)
        chart_swr_out.set_categories(cats)
        ws.add_chart(chart_swr_out, 'B18')

        chart_phase = openpyxl.chart.LineChart()
        chart_phase.style = 2
        chart_phase.x_axis.title = 'F, GHz'
        chart_phase.y_axis.title = 'Phase, deg'
        vals = openpyxl.chart.Reference(ws, range_string=f'Sheet!$BT$1:$CH${max_row}')
        chart_phase.add_data(vals, titles_from_data=True)
        chart_phase.set_categories(cats)
        ws.add_chart(chart_phase, 'L18')

        chart_phase_err = openpyxl.chart.LineChart()
        chart_phase_err.style = 2
        chart_phase_err.x_axis.title = 'F, GHz'
        chart_phase_err.y_axis.title = 'Phase err, deg'
        vals = openpyxl.chart.Reference(ws, range_string=f'Sheet!$CI$1:$CW${max_row}')
        chart_phase_err.add_data(vals, titles_from_data=True)
        chart_phase_err.set_categories(cats)
        ws.add_chart(chart_phase_err, 'B34')

        stats_cell = ws['L34']
        stats_cell.value = f'delta Kp'
        stats_cell.offset(0, 1).value = f'{self._result._delta_Kp:.02f}'
        stats_cell.offset(1, 0).value = f'SWR in max'
        stats_cell.offset(1, 1).value = f'{self._result._swr_in_max:.01f}'
        stats_cell.offset(2, 0).value = f'SWR out max'
        stats_cell.offset(2, 1).value = f'{self._result._swr_out_max:.01f}'
        stats_cell.offset(3, 0).value = f'phase err'
        stats_cell.offset(3, 1).value = f'{self._result._phase_err_min:.01f}..{self._result._phase_err_max:.01f}'
        stats_cell.offset(4, 0).value = f'S21 max'
        stats_cell.offset(4, 1).value = f'{self._result._s21_max_in_range:.02f}'
        stats_cell.offset(5, 0).value = f'S21 delta max'
        stats_cell.offset(5, 1).value = f'{self._result._s21_delta_max:.02f}'
        stats_cell.offset(6, 0).value = f'S21 deltas'

        delas_cell = stats_cell.offset(7, 1)
        for i, state in enumerate(self._result._states):
            delas_cell.offset(i, 0).value = f'{state:.01f}'
        for i, delta in enumerate(self._result._s21_deltas):
            delas_cell.offset(i, 1).value = f'{delta:.02f}'

        wb.save(self.file_path)
        print('\nsaved .xlsx:', self.file_path)

        subprocess.call('explorer ' + '.\\xlsx\\', shell=True)
